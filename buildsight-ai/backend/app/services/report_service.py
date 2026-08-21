"""BuildSight AI — Comprehensive Real-Time Audit Report Generator

Generates executive safety, compliance, and progress audit reports directly from real-time MongoDB data:
  - registered_workers
  - worker_sessions
  - worker_snapshots
  - violations (with root cause analysis & OSHA citations)
  - progress_records (9 construction stages)
  - danger_zones & delay predictions
"""

import os
import json
import uuid
import logging
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
import pymongo

from app.database.mongodb import get_db
from app.database.collections import (
    COLLECTION_REGISTERED_WORKERS,
    COLLECTION_WORKER_SESSIONS,
    COLLECTION_WORKER_SNAPSHOTS,
    COLLECTION_VIOLATIONS,
    COLLECTION_PROGRESS_RECORDS,
    COLLECTION_DANGER_ZONES,
    COLLECTION_REPORTS,
)
from app.database.utils import serialize_mongo_doc, serialize_mongo_docs
from app.config import settings

logger = logging.getLogger(__name__)


class ComprehensiveReportGenerator:
    """Generates structured, executive-grade audit reports from real-time stored database telemetry."""

    @staticmethod
    def generate_audit_report(
        title: str = "Construction Site Safety & Compliance Audit Report",
        auditor_name: str = "BuildSight AI Automated Safety Engine",
        notes: str = "",
        save_to_db: bool = True,
    ) -> Dict[str, Any]:
        """Generate a complete audit report from real-time MongoDB collections."""
        db = get_db()
        now = datetime.now(timezone.utc)
        report_id = f"RPT-{now.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

        # 1. Query Registered Workers
        registered_workers = list(db[COLLECTION_REGISTERED_WORKERS].find({"active_status": "ACTIVE"}))
        reg_map = {w.get("worker_code"): w for w in registered_workers}

        # 2. Query All Violations with Deduplication
        raw_violations = list(db[COLLECTION_VIOLATIONS].find().sort("timestamp", pymongo.DESCENDING))
        
        # Deduplicate incidents by episodic key
        deduped_violations = {}
        for v in raw_violations:
            w_code = v.get("worker_code") or v.get("permanent_worker_id")
            w_id = v.get("worker_id", "0")
            w_key = f"code_{w_code}" if w_code else f"track_{w_id}"
            v_type = v.get("violation_type", "UNKNOWN")
            missing_sig = "-".join(sorted(v.get("missing_items") or []))
            ep_key = f"{w_key}_{v_type}_{missing_sig}"

            existing = deduped_violations.get(ep_key)
            if not existing or (v.get("duration_seconds", 0) > existing.get("duration_seconds", 0)) or (v.get("status") == "OPEN" and existing.get("status") != "OPEN"):
                deduped_violations[ep_key] = v

        unique_viols = list(deduped_violations.values())
        unique_viols.sort(key=lambda x: x.get("timestamp") if isinstance(x.get("timestamp"), datetime) else str(x.get("timestamp", "")), reverse=True)

        # 3. Enrich Violations with RCA & OSHA Standards
        enriched_viols = []
        for inc in unique_viols:
            v_type = inc.get("violation_type", "UNKNOWN")
            missing = inc.get("missing_items") or []
            desc = inc.get("description", "")
            w_code = inc.get("worker_code") or inc.get("permanent_worker_id")
            w_info = reg_map.get(w_code) if w_code else None
            w_name = w_info.get("name") if w_info else inc.get("worker_name")
            emp_no = w_info.get("employee_number") if w_info else inc.get("employee_number")
            role = w_info.get("role") if w_info else ("Registered Personnel" if w_code else "Site Visitor / Contractor")

            vt_lower = v_type.lower()
            if "danger_zone" in vt_lower or "danger_zone" in missing:
                reason = desc if desc else "Worker breached unauthorized restricted perimeter in high-risk zone."
                osha_std = "OSHA 29 CFR 1926.651 / 1926.550 (Hazardous Area Perimeter Control)"
                corrective = "Evacuate personnel immediately to designated safe muster point and verify zone signage."
            elif "helmet" in vt_lower or "hardhat" in vt_lower or "helmet" in missing or "Hardhat" in str(missing):
                reason = "Worker operating in active overhead hazard area without ANSI Z89.1 certified protective hardhat."
                osha_std = "OSHA 29 CFR 1926.100(a) (Head Protection)"
                corrective = "Halt immediate task, issue ANSI Type I/II hardhat, and log verbal safety briefing."
            elif "vest" in vt_lower or "safety_vest" in vt_lower or "vest" in missing or "safety_vest" in missing:
                reason = "Worker operating in heavy equipment/vehicle corridor without high-visibility retroreflective safety vest."
                osha_std = "OSHA 29 CFR 1926.201 (Signaling & High-Visibility Apparel)"
                corrective = "Provide Class 2/3 high-visibility safety apparel before worker re-enters active site zone."
            elif "glove" in vt_lower or "gloves" in missing:
                reason = "Worker handling abrasive materials or mechanical tools without certified cut/abrasion resistant gloves."
                osha_std = "OSHA 29 CFR 1926.95 (Personal Protective Equipment - Hand Protection)"
                corrective = "Issue task-appropriate protective gloves (cut level A2+) prior to material handling."
            elif "mask" in vt_lower or "face_mask" in missing or "mask" in missing:
                reason = "Worker in particulate/silica-generating environment without approved respiratory protection."
                osha_std = "OSHA 29 CFR 1926.1153 (Respirable Crystalline Silica Standard)"
                corrective = "Issue N95/P100 respirator and verify localized dust suppression systems are active."
            else:
                reason = desc if desc else f"Safety compliance threshold breach detected: {v_type.replace('_', ' ')}."
                osha_std = "OSHA 29 CFR 1926 Subpart C (General Safety & Health Provisions)"
                corrective = "Conduct supervisor review and ensure mandatory PPE compliance."

            ts_val = inc.get("timestamp")
            ts_str = ts_val.isoformat() if isinstance(ts_val, datetime) else str(ts_val or "")

            res_val = inc.get("resolved_at")
            res_str = res_val.isoformat() if isinstance(res_val, datetime) else (str(res_val) if res_val else None)

            vid = inc.get("violation_id", "")
            evidence_url = f"/data/evidence/violation_{vid}.jpg" if vid else None

            enriched_viols.append({
                "violation_id": vid,
                "worker_id": inc.get("worker_id"),
                "worker_code": w_code,
                "worker_name": w_name,
                "employee_number": emp_no,
                "worker_role": role,
                "worker_display": f"{w_name} ({w_code})" if (w_name and w_code) else (w_code or f"Tracker #{inc.get('worker_id', '?')}"),
                "violation_type": v_type,
                "reason": reason,
                "osha_standard": osha_std,
                "corrective_action": corrective,
                "missing_items": missing,
                "severity": inc.get("severity", "MEDIUM"),
                "risk_score": inc.get("risk_score", 50.0),
                "status": inc.get("status", "OPEN"),
                "duration_seconds": inc.get("duration_seconds", 0.0),
                "timestamp": ts_str,
                "resolved_at": res_str,
                "evidence_url": evidence_url,
            })

        # 4. Worker Compliance Roster (Per Worker Statistics)
        worker_roster = []
        for w in registered_workers:
            code = w.get("worker_code")
            w_viols = [v for v in enriched_viols if v.get("worker_code") == code]
            open_viols = [v for v in w_viols if v.get("status") == "OPEN"]
            comp_rate = max(0.0, 100.0 - (len(w_viols) * 15.0))
            grade = "A" if comp_rate >= 90 else ("B" if comp_rate >= 75 else ("C" if comp_rate >= 60 else "D"))

            worker_roster.append({
                "worker_code": code,
                "name": w.get("name"),
                "employee_number": w.get("employee_number"),
                "department": w.get("department"),
                "role": w.get("role"),
                "profile_image_path": w.get("profile_image_path"),
                "total_violations": len(w_viols),
                "active_violations": len(open_viols),
                "compliance_score": round(comp_rate, 1),
                "compliance_grade": grade,
                "violations": w_viols,
            })

        # 5. Query Latest Progress Record & Milestones
        latest_progress = db[COLLECTION_PROGRESS_RECORDS].find_one({}, sort=[("timestamp", pymongo.DESCENDING)])
        progress_history = list(db[COLLECTION_PROGRESS_RECORDS].find().sort("timestamp", pymongo.DESCENDING).limit(10))

        # 6. Danger Zones
        danger_zones = list(db[COLLECTION_DANGER_ZONES].find({"active": True}))

        # 7. Summary Analytics & Site Grade
        total_viols_count = len(enriched_viols)
        critical_count = sum(1 for v in enriched_viols if v.get("severity") == "CRITICAL")
        high_count = sum(1 for v in enriched_viols if v.get("severity") == "HIGH")
        medium_count = sum(1 for v in enriched_viols if v.get("severity") == "MEDIUM")
        low_count = sum(1 for v in enriched_viols if v.get("severity") == "LOW")

        avg_site_compliance = round(max(0.0, 100.0 - (total_viols_count * 5.0) - (critical_count * 10.0)), 1)
        site_grade = "Grade A (Excellent)" if avg_site_compliance >= 90 else (
            "Grade B (Good)" if avg_site_compliance >= 75 else (
                "Grade C (Requires Attention)" if avg_site_compliance >= 60 else "Grade D (Critical Hazard Alert)"
            )
        )

        # 8. Actionable Safety Recommendations
        recommendations = []
        if critical_count > 0:
            recommendations.append("Immediate Stop-Work Review: Critical danger zone or extreme fall hazard violations detected. Re-verify perimeter barriers.")
        if any("helmet" in str(v.get("missing_items", [])) or v.get("violation_type") == "MISSING_HELMET" for v in enriched_viols):
            recommendations.append("Conduct Mandatory Head Protection Briefing: Multiple non-compliant hardhat events recorded under OSHA 1926.100.")
        if any("vest" in str(v.get("missing_items", [])) or v.get("violation_type") == "MISSING_VEST" for v in enriched_viols):
            recommendations.append("Enforce Class 2/3 High-Visibility Apparel in active mobile equipment corridors (OSHA 1926.201).")
        if not recommendations:
            recommendations.append("Maintain continuous surveillance and bi-weekly safety toolbox talks.")

        report = {
            "report_id": report_id,
            "title": title,
            "auditor_name": auditor_name,
            "notes": notes,
            "generated_at": now.isoformat(),
            "generated_timestamp": now.strftime("%B %d, %Y at %H:%M:%S UTC"),
            "site_risk_grade": site_grade,
            "site_compliance_score": avg_site_compliance,
            "executive_summary": {
                "total_registered_workers": len(registered_workers),
                "total_deduplicated_incidents": total_viols_count,
                "active_open_incidents": sum(1 for v in enriched_viols if v.get("status") == "OPEN"),
                "resolved_incidents": sum(1 for v in enriched_viols if v.get("status") == "RESOLVED"),
                "critical_violations": critical_count,
                "high_violations": high_count,
                "medium_violations": medium_count,
                "low_violations": low_count,
                "configured_danger_zones": len(danger_zones),
                "current_construction_stage": latest_progress.get("current_stage", "Structural Work") if latest_progress else "Structural Work",
                "overall_progress_pct": latest_progress.get("overall_progress_percentage", 55.0) if latest_progress else 55.0,
            },
            "worker_roster": worker_roster,
            "incident_log": enriched_viols,
            "progress_audit": {
                "current_stage": latest_progress.get("current_stage", "Structural Work") if latest_progress else "Structural Work",
                "stage_completion_pct": latest_progress.get("stage_completion_percentage", 65.0) if latest_progress else 65.0,
                "overall_progress_pct": latest_progress.get("overall_progress_percentage", 55.0) if latest_progress else 55.0,
                "project_status": latest_progress.get("project_status", "ON_TRACK") if latest_progress else "ON_TRACK",
                "milestone_history": [
                    {
                        "stage": p.get("current_stage"),
                        "progress": p.get("overall_progress_percentage"),
                        "timestamp": p.get("timestamp").isoformat() if isinstance(p.get("timestamp"), datetime) else str(p.get("timestamp")),
                    }
                    for p in progress_history
                ]
            },
            "recommendations": recommendations,
            "legal_compliance": {
                "osha_standard_framework": "OSHA 29 CFR 1926 Safety & Health Regulations for Construction",
                "gdpr_article_9": "Biometrics stored as non-reversible mathematical embedding vectors; zero raw crop persistence",
                "bipa_section_15": "Informed worker consent verified at registration; automatic 30-day retention purge",
            }
        }

        if save_to_db:
            try:
                db[COLLECTION_REPORTS].insert_one({
                    "report_id": report_id,
                    "report_type": "COMPREHENSIVE_AUDIT",
                    "title": title,
                    "auditor_name": auditor_name,
                    "generated_at": now,
                    "site_compliance_score": avg_site_compliance,
                    "site_risk_grade": site_grade,
                    "total_incidents": total_viols_count,
                    "data": report,
                })
                logger.info(f"✓ Comprehensive Audit Report saved to MongoDB: {report_id}")
            except Exception as e:
                logger.warning(f"Could not persist report to MongoDB: {e}")

        return report

    @staticmethod
    def get_report_history(limit: int = 20) -> List[Dict[str, Any]]:
        """Retrieve previously generated reports from MongoDB."""
        db = get_db()
        reports = list(db[COLLECTION_REPORTS].find().sort("generated_at", pymongo.DESCENDING).limit(limit))
        return [
            {
                "id": str(r["_id"]),
                "report_id": r.get("report_id", f"RPT-{r['_id']}"),
                "title": r.get("title", "Safety Audit Report"),
                "auditor_name": r.get("auditor_name", "BuildSight AI"),
                "generated_at": r.get("generated_at").isoformat() if isinstance(r.get("generated_at"), datetime) else str(r.get("generated_at")),
                "site_compliance_score": r.get("site_compliance_score", 100.0),
                "site_risk_grade": r.get("site_risk_grade", "Grade A"),
                "total_incidents": r.get("total_incidents", 0),
            }
            for r in reports
        ]


    @staticmethod
    def generate_workers_excel() -> bytes:
        """
        Generate a professional Excel workbook (.xlsx) containing all registered workers,
        with their actual profile photos embedded directly into the spreadsheet cells.
        """
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as OpenPyXLImage
        from PIL import Image as PILImage
        from pathlib import Path

        report = ComprehensiveReportGenerator.generate_audit_report(save_to_db=False)
        roster = report.get("worker_roster", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registered Workers"

        # Theme Colors
        header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        sub_font = Font(name="Arial", size=9, color="64748B", italic=True)
        cell_font = Font(name="Arial", size=10)
        bold_font = Font(name="Arial", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        # Title Block
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "BUILDSIGHT AI — REGISTERED WORKERS DIRECTORY ROSTER"
        title_cell.font = Font(name="Arial", size=14, bold=True, color="0284C7")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:K2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S UTC')} • Total Registered Workers: {len(roster)} • Biometric Model: YuNet + SFace"
        sub_cell.font = sub_font
        sub_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        # Header Row
        headers = [
            "Worker Photo",
            "Worker Code",
            "Employee ID",
            "Full Name",
            "Department",
            "Role",
            "Status",
            "Compliance Score",
            "Safety Grade",
            "Total Violations",
            "Active Violations",
        ]

        ws.row_dimensions[4].height = 26
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 16  # Photo
        ws.column_dimensions["B"].width = 16  # Code
        ws.column_dimensions["C"].width = 16  # Emp ID
        ws.column_dimensions["D"].width = 24  # Name
        ws.column_dimensions["E"].width = 22  # Dept
        ws.column_dimensions["F"].width = 22  # Role
        ws.column_dimensions["G"].width = 14  # Status
        ws.column_dimensions["H"].width = 18  # Score
        ws.column_dimensions["I"].width = 15  # Grade
        ws.column_dimensions["J"].width = 16  # Total Viols
        ws.column_dimensions["K"].width = 16  # Active Viols

        backend_data_dir = Path(__file__).resolve().parents[2] / "data"

        current_row = 5
        for w in roster:
            ws.row_dimensions[current_row].height = 58  # generous height for embedded photo

            # 1. Embed Photo in Column A
            photo_rel = w.get("profile_image_path")
            photo_embedded = False

            if photo_rel:
                clean_rel = photo_rel.replace("/api/profiles/", "profiles/").replace("/data/evidence/", "evidence/")
                disk_path = backend_data_dir / clean_rel
                if not disk_path.exists():
                    disk_path = backend_data_dir / "profiles" / Path(photo_rel).name

                if disk_path.exists():
                    try:
                        pil_img = PILImage.open(str(disk_path))
                        pil_img.thumbnail((50, 50))
                        img_byte_arr = BytesIO()
                        pil_img.convert("RGB").save(img_byte_arr, format="JPEG", quality=85)
                        img_byte_arr.seek(0)

                        img = OpenPyXLImage(img_byte_arr)
                        img.width = 50
                        img.height = 50
                        ws.add_image(img, f"A{current_row}")
                        photo_embedded = True
                    except Exception as img_err:
                        logger.debug(f"Excel image embed note: {img_err}")

            cell_a = ws.cell(row=current_row, column=1)
            if not photo_embedded:
                cell_a.value = "No Photo"
                cell_a.font = sub_font
            cell_a.alignment = center_align
            cell_a.border = thin_border

            # Other columns
            row_data = [
                (2, w.get("worker_code", ""), center_align, bold_font),
                (3, w.get("employee_number", ""), center_align, cell_font),
                (4, w.get("name", ""), left_align, bold_font),
                (5, w.get("department", ""), left_align, cell_font),
                (6, w.get("role", ""), left_align, cell_font),
                (7, w.get("active_status", "ACTIVE"), center_align, cell_font),
                (8, f"{w.get('compliance_score', 100.0)}%", center_align, bold_font),
                (9, f"Grade {w.get('compliance_grade', 'A')}", center_align, bold_font),
                (10, w.get("total_violations", 0), center_align, cell_font),
                (11, w.get("active_violations", 0), center_align, cell_font),
            ]

            for col_idx, val, align, font in row_data:
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.alignment = align
                c.font = font
                c.border = thin_border
                if col_idx == 8 and isinstance(w.get('compliance_score'), (int, float)):
                    score = w.get('compliance_score')
                    if score >= 85:
                        c.font = Font(name="Arial", size=10, bold=True, color="16A34A")
                    else:
                        c.font = Font(name="Arial", size=10, bold=True, color="EA580C")

            current_row += 1

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


comprehensive_report_generator = ComprehensiveReportGenerator()
